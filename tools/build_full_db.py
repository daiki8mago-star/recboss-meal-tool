"""
build_full_db.py  (PFC対応版)
文部科学省『日本食品標準成分表（八訂）増補2023年』の公式Excelから
全食品（約2,478件）の food_calories_full.json を一括生成する。

▼ 元のスクリプトはエネルギー(kcal)しか取り出していなかったが、
   この版は たんぱく質 / 脂質 / 炭水化物 も取り出す。
   → 食事プラン作成ツールが必要とする PFC をそのまま使えるDBになる。

【手順】
1. 下記ページから「食品成分表(本表) Excel」をダウンロード:
   https://www.mext.go.jp/a_menu/syokuhinseibun/mext_00001.html
   （ファイル名例: 20230428-mxt_kagsei-mext_00001_011.xlsx）
2. このスクリプトと同じフォルダに置く
3. pip install openpyxl
4. まず列位置の確認:  python build_full_db.py <Excel> --inspect
   → 先頭数行が表示されるので、食品名/エネルギー/たんぱく質/脂質/炭水化物の
     列番号(0始まり)を確認する
5. 生成:  python build_full_db.py <Excel> \
            --name 3 --energy 5 --protein 6 --fat 8 --carb 11

※ 列番号は版によりずれるため、必ず --inspect で確認してから指定すること。
   指定しない場合は八訂2023での代表値（下の DEFAULTS）を使う。
"""

import sys
import json
import argparse

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl が必要です:  pip install openpyxl")

# 八訂2023 本表でよく合う既定の列位置（0始まり）。版が違えば --inspect で確認のこと。
DEFAULTS = {"name": 3, "energy": 5, "protein": 6, "fat": 8, "carb": 11, "start": 12}


def _num(v):
    """'Tr'・'-'・'(0)'・空欄などを 0.0 に、数値は float に。失敗は None。"""
    if v in (None, "", "-", "Tr", "tr"):
        return 0.0
    s = str(v).replace("(", "").replace(")", "").replace(",", "").strip()
    if s in ("", "-", "Tr", "tr"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return None


def inspect(xlsx_path):
    ws = openpyxl.load_workbook(xlsx_path, data_only=True).active
    print(f"シート: {ws.title}  最大列: {ws.max_column}  最大行: {ws.max_row}\n")
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=16, values_only=True)):
        cells = " | ".join(f"[{i}]{('' if v is None else str(v))[:10]}" for i, v in enumerate(row))
        print(f"行{r+1}: {cells}")


def build(xlsx_path, cols, out_path="food_calories_full.json"):
    ws = openpyxl.load_workbook(xlsx_path, data_only=True).active
    foods, skipped = {}, 0
    for row in ws.iter_rows(min_row=cols["start"], values_only=True):
        def cell(key):
            i = cols[key]
            return row[i] if len(row) > i else None
        name = cell("name")
        energy = _num(cell("energy"))
        if not name or energy is None or energy == 0.0:
            skipped += 1
            continue
        p, f, c = _num(cell("protein")), _num(cell("fat")), _num(cell("carb"))
        foods[str(name).strip()] = {
            "kcal_per_100g": round(energy, 1),
            "protein_g": round(p or 0, 1),
            "fat_g": round(f or 0, 1),
            "carb_g": round(c or 0, 1),
        }
    out = {
        "meta": {
            "source": "日本食品標準成分表（八訂）増補2023年（公式Excelから自動生成・PFC対応）",
            "unit": "100gあたり（可食部）",
            "count": len(foods),
        },
        "foods": foods,
    }
    with open(out_path, "w", encoding="utf-8") as fp:
        json.dump(out, fp, ensure_ascii=False, indent=2)
    print(f"{len(foods)} 食品を書き出しました（スキップ {skipped} 行）-> {out_path}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="公式Excelのパス")
    ap.add_argument("--inspect", action="store_true", help="先頭行を表示して列位置を確認")
    for k in ("name", "energy", "protein", "fat", "carb", "start"):
        ap.add_argument(f"--{k}", type=int, default=DEFAULTS[k])
    args = ap.parse_args()

    if args.inspect:
        inspect(args.xlsx)
    else:
        cols = {k: getattr(args, k) for k in ("name", "energy", "protein", "fat", "carb", "start")}
        build(args.xlsx, cols)
